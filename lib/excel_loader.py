import csv
import os
import pandas as pd
import json
import openpyxl as openpyxl
import config.settings as settings
import hashlib
import lib.common_utils as common_utils


ROOT_PATH = 'data/'
GENERAL_INFO_PATH = os.path.join('data', 'general_info.xlsx')
RISK_QUESTIONS_PATH = os.path.join('data', 'risk_characteristics.xlsx')


def hash_text(text):
    m = hashlib.md5(text.encode('UTF-8'))
    return m.hexdigest()


def format_text_field(value):
    return str(value).lower().strip()


class ExceLoader():
    def __init__(self):
        cache_file = os.path.join(ROOT_PATH + 'cache.json')
        disable_cache = False
        # disable_cache = True

        if not disable_cache and os.path.isfile(cache_file):
            print('loading cache')
            with open(cache_file, "r") as jsonfile:
                dump_obj = json.load(jsonfile)

        else:
            with open(cache_file, "w") as outfile:
                print('loading excel files')
                # need the names
                self.valueweb = self._valueweb_lookup()
                dump_obj = {
                    'valueweb': self.valueweb,
                    'risk_questions': self._risk_questions(),
                    'risk_question_sections': self._risk_question_sections(),
                    'goal_risk_defaults': self._goal_risk_defaults(),
                    'be_questions': self._be_questions(),
                    'be_lookup': self._be_lookup(),
                }
                json.dump(dump_obj, outfile)

        self.risk_questions = dump_obj['risk_questions']
        self.risk_question_sections = dump_obj['risk_question_sections']
        self.goal_risk_defaults = dump_obj['goal_risk_defaults']
        self.valueweb = dump_obj['valueweb']
        self.be_questions = dump_obj['be_questions']
        self.be_lookup = dump_obj['be_lookup']

    def _risk_questions(self):
        product_questions = {}
        non_product_questions = {}
        file = ROOT_PATH + 'risk_characteristics.xlsx'
        wb = openpyxl.load_workbook(filename=file)
        total = 0
        codes = []
        start_tab = 2

        question_tabs = {
            'business_inputs': wb[wb.sheetnames[2]],
            'operations': wb[wb.sheetnames[3]],
            'employees': wb[wb.sheetnames[4]],
            'products_and_customers': wb[wb.sheetnames[5]],
            'broader_society': wb[wb.sheetnames[6]],
        }

        default_risks_tab = wb[wb.sheetnames[7]]

        counts = {}
        question_ids_by_goal = {}

        for key in question_tabs.keys():
            counts[key] = {}
            product_questions[key] = {'name': self.valueweb[key]['name'], 'items': {}, }
            non_product_questions[key] = {'name': self.valueweb[key]['name'], 'items': {}, }

            tab = question_tabs[key]
            rows = 0

            for row in tab.iter_rows(max_col=10, min_row=2, max_row=50):
                code = row[0].value
                new_group_title = row[1].value

                if code and code in codes:
                    print("!!!! DUPLICATE ID/CODE IN RISK ASSESSMENT - PLS FIX")
                    print(code, row)
                    print('skipping')
                    continue

                if code:
                    code = str(code)
                    codes.append(code)
                    rows += 1
                    risk = row[4].value.strip()
                    if risk not in ['High', 'Low', 'Unlikely', 'Moderate']:
                        print('warning - risk not found', id)

                    goals = row[5].value
                    if goals:
                        goals = goals.split(',')
                        goals = [goal.strip() for goal in goals]

                    product = row[6].value == 'Yes'
                    non_product = row[7].value == 'Yes'

                    question = {
                        'order': rows,
                        'code': code,
                        'title': row[2].value,
                        'help_text': row[3].value,
                        'risk': risk,
                        'goals': goals,
                        'number': rows,
                    }

                    if product:
                        product_questions[key]['items'][code] = question
                        for goal in goals:
                            if goal not in question_ids_by_goal:
                                question_ids_by_goal[goal] = {'product': [], 'non_product': []}
                            question_ids_by_goal[goal]['product'].append(code)

                    if non_product:
                        non_product_questions[key]['items'][str(code)] = question
                        for goal in goals:
                            if goal not in question_ids_by_goal:
                                question_ids_by_goal[goal] = {'product': [], 'non_product': []}
                            question_ids_by_goal[goal]['non_product'].append(code)

            total += rows
            counts[key] = {
                'product': len(product_questions[key]['items']),
                'non_product': len(non_product_questions[key]['items'])
            }

        for goal in question_ids_by_goal:
            if not question_ids_by_goal[goal]:

                print('no goals', goal)

        return {
            'product': product_questions,
            'non_product': non_product_questions,
            'counts': counts,
            'question_ids_by_goal': question_ids_by_goal,
        }

    def _risk_question_sections(self):
        """
        inject after
        product type > valueweb section > question code > group_title
        """
        product_sections = {}
        non_product_sections = {}
        wb = openpyxl.load_workbook(filename=RISK_QUESTIONS_PATH)

        tabs = {
            'business_inputs': wb[wb.sheetnames[2]],
            'operations': wb[wb.sheetnames[3]],
            'employees': wb[wb.sheetnames[4]],
            'products_and_customers': wb[wb.sheetnames[5]],
            'broader_society': wb[wb.sheetnames[6]],
        }

        for key in tabs.keys():
            product_sections[key] = {}
            non_product_sections[key] = {}
            tab = tabs[key]

            current = ''
            p_section = ''
            np_section = ''

            for row in tab.iter_rows(max_col=10, min_row=2, max_row=90):
                code = row[0].value
                new_group_title = row[1].value

                if new_group_title and new_group_title != current:
                    current = new_group_title
                    p_section = new_group_title
                    np_section = new_group_title

                if code is not None:
                    print(code, new_group_title)
                    # print(code)
                    code = str(code)
                    product = row[6].value == 'Yes'
                    non_product = row[7].value == 'Yes'

                    if product and p_section:
                        product_sections[key][code] = p_section
                        p_section = ''

                    if non_product and np_section:
                        non_product_sections[key][code] = np_section
                        np_section = ''

        # print(product_sections)
        return {
            'product': product_sections,
            'non_product': non_product_sections,
        }

    def _goal_risk_defaults(self):
        risks = {}
        risks_defaults_tab = 8
        risks_defaults_descriptions_tab = 7
        df = pd.read_excel(ROOT_PATH + 'risk_characteristics.xlsx',
                           sheet_name=[risks_defaults_tab, risks_defaults_descriptions_tab], skiprows=1)

        for row in df[risks_defaults_tab].values:
            code = row[0]
            if code and isinstance(code, str):
                code = code.strip()
                product_risk = ''
                non_product_risk = ''
                default = 'default'

                if format_text_field(row[1]) == default:
                    product_risk = common_utils.HIGH_RISK
                elif format_text_field(row[2]) == default:
                    product_risk = common_utils.MODERATE_RISK
                elif format_text_field(row[3]) == default:
                    product_risk = common_utils.LOW_RISK
                elif format_text_field(row[4]) == default:
                    product_risk = common_utils.UNLIKELY_RISK

                if format_text_field(row[7]) == default:
                    non_product_risk = common_utils.HIGH_RISK
                elif format_text_field(row[8]) == default:
                    non_product_risk = common_utils.MODERATE_RISK
                elif format_text_field(row[9]) == default:
                    non_product_risk = common_utils.LOW_RISK
                elif format_text_field(row[10]) == default:
                    non_product_risk = common_utils.UNLIKELY_RISK

                risks[code] = {
                    'product': product_risk,
                    'non_product': non_product_risk,
                }

        for row in df[risks_defaults_descriptions_tab].values:
            code = row[0]
            if code and isinstance(code, str):
                risks[code].update({
                    'product_title': row[2],
                    'non_product_title': row[4],
                })

        return risks

    def _valueweb_lookup(self):
        questions = {}

        wb = openpyxl.load_workbook(filename=GENERAL_INFO_PATH)
        tab = wb[wb.sheetnames[2]]
        lookup = {}
        order = 0

        for row in tab.iter_rows(max_col=10, min_row=2, max_row=10):
            name = row[0].value
            order += 1

            if name:
                key = name.lower().replace(' ', '_')
                lookup[key] = {
                    'name': name,
                    'order': order,
                    'goals': sorted((row[1].value).split(',')),
                    'description': row[2].value,
                    'description_2': row[4].value,
                    'description_short': row[3].value,
                }
        return lookup

    def _be_lookup(self):
        questions = {}
        wb = openpyxl.load_workbook(filename=GENERAL_INFO_PATH)
        tab = wb[wb.sheetnames[1]]
        lookup = {}
        order = 0

        for row in tab.iter_rows(max_col=10, min_row=2, max_row=30):
            name = row[0].value
            order += 1

            if name:
                lookup[name] = {
                    'name': name,
                    'group': row[1].value,
                    'input_type': row[2].value,
                    'input_name': row[3].value,
                    'short_name': row[8].value,
                }
        return lookup

    def _be_questions(self):
        questions = {}
        wb = openpyxl.load_workbook(filename=GENERAL_INFO_PATH)
        first_tab = wb[wb.sheetnames[0]]
        current_goal = ''
        current_sub_type = ''
        goal_count = 0

        for row in first_tab.iter_rows(max_col=10, min_row=2):
            goal = row[0].value
            sub_type = row[1].value
            group_title = row[2].value
            input_type = row[3].value
            input_unit = row[4].value
            text = row[5].value

            if goal and goal != current_goal:
                current_goal = goal
                current_sub_type = ''

            if sub_type and sub_type != current_sub_type:
                current_sub_type = sub_type

            if text:
                model = hash_text(current_goal + input_type + text)
                obj = {
                    'text': text,
                    'type': input_type,
                    'model': model,
                    'unit': input_unit,
                }
                if current_goal not in questions:
                    questions[current_goal] = {}

                if current_sub_type:
                    if current_sub_type not in questions[current_goal]:
                        questions[current_goal][current_sub_type] = {}
                    questions[current_goal][current_sub_type][model] = obj
                else:
                    questions[current_goal][model] = obj

        return questions
